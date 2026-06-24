"""Branded XLSX renderer used by the report endpoints."""

from __future__ import annotations

import io
from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.config import settings


# PUG palette
NAVY = "FF1A234A"
GOLD = "FFC9A14A"
LIGHT = "FFF7F8FC"
BORDER = "FFE3E7F0"


def render_xlsx(
    *,
    title: str,
    subtitle: str,
    columns: list[dict[str, Any]],
    rows: list[dict[str, Any]],
    params: dict[str, Any] | None = None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:30] or "Report"

    n_cols = len(columns)
    col_letter_last = get_column_letter(max(n_cols, 1))

    # ---- Brand title row ----
    ws.merge_cells(f"A1:{col_letter_last}1")
    cell = ws["A1"]
    cell.value = f"{settings.brand_company_name}  -  Legal Case Control System"
    cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFFFF")
    cell.fill = PatternFill("solid", fgColor=NAVY)
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    # Gold accent row
    ws.merge_cells(f"A2:{col_letter_last}2")
    accent = ws["A2"]
    accent.fill = PatternFill("solid", fgColor=GOLD)
    ws.row_dimensions[2].height = 4

    # ---- Report title row ----
    ws.merge_cells(f"A3:{col_letter_last}3")
    t = ws["A3"]
    t.value = title
    t.font = Font(name="Calibri", size=13, bold=True, color="FF1A234A")
    t.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[3].height = 22

    # ---- Subtitle / parameters row ----
    sub_parts: list[str] = []
    if subtitle:
        sub_parts.append(subtitle)
    if params:
        non_empty = {k: v for k, v in params.items() if v not in (None, "")}
        if non_empty:
            sub_parts.append(" | ".join(f"{k}: {v}" for k, v in non_empty.items()))
    sub_parts.append(f"Generated: {datetime.utcnow():%Y-%m-%d %H:%M UTC}")
    ws.merge_cells(f"A4:{col_letter_last}4")
    sub = ws["A4"]
    sub.value = " - ".join(sub_parts)
    sub.font = Font(name="Calibri", size=10, italic=True, color="FF5B6478")
    sub.alignment = Alignment(horizontal="left", indent=1)

    # ---- Header row ----
    header_row = 6
    border_thin = Border(
        bottom=Side(style="thin", color=BORDER),
        right=Side(style="thin", color=BORDER),
    )
    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=header_row, column=i, value=str(col["label"]))
        c.font = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        c.border = border_thin
    ws.row_dimensions[header_row].height = 22
    ws.freeze_panes = f"A{header_row + 1}"

    # ---- Data rows ----
    for r_offset, row in enumerate(rows, start=header_row + 1):
        for i, col in enumerate(columns, start=1):
            key = col["key"]
            ctype = col.get("type", "text")
            val = row.get(key)
            cell = ws.cell(row=r_offset, column=i, value=_coerce(val, ctype))
            cell.border = border_thin
            cell.font = Font(name="Calibri", size=10, color="FF1C2233")
            if ctype == "number":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right", indent=1)
            elif ctype == "int":
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right", indent=1)
            elif ctype == "date":
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = Alignment(horizontal="left", indent=1)
            elif ctype == "datetime":
                cell.number_format = "yyyy-mm-dd hh:mm"
                cell.alignment = Alignment(horizontal="left", indent=1)
            else:
                cell.alignment = Alignment(horizontal="left", indent=1, wrap_text=False)
        # Zebra stripe
        if r_offset % 2 == 0:
            for i in range(1, n_cols + 1):
                ws.cell(row=r_offset, column=i).fill = PatternFill("solid", fgColor=LIGHT)

    # ---- Auto-ish column widths ----
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        max_len = len(str(col["label"]))
        for r in range(header_row + 1, header_row + 1 + len(rows)):
            v = ws.cell(row=r, column=i).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 50)

    # ---- Footer ----
    foot_row = header_row + 1 + len(rows) + 1
    ws.merge_cells(f"A{foot_row}:{col_letter_last}{foot_row}")
    f = ws.cell(row=foot_row, column=1)
    f.value = f"{len(rows)} row(s)  -  (c) {settings.brand_company_name}"
    f.font = Font(name="Calibri", size=9, italic=True, color="FF8A91A3")
    f.alignment = Alignment(horizontal="left", indent=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _coerce(v: Any, ctype: str) -> Any:
    if v is None:
        return None
    if ctype in ("number",):
        if isinstance(v, Decimal):
            return float(v)
        try:
            return float(v)
        except (TypeError, ValueError):
            return v
    if ctype == "int":
        try:
            return int(v)
        except (TypeError, ValueError):
            return v
    if ctype == "date" and isinstance(v, str):
        try:
            return date.fromisoformat(v)
        except ValueError:
            return v
    if ctype == "datetime" and isinstance(v, str):
        try:
            return datetime.fromisoformat(v)
        except ValueError:
            return v
    return v
